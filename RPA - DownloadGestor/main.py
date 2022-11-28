import time
import uiautomation as uia
from selenium.webdriver.common.by import By
import win32gui
import win32con
import win32api
from datetime import date
from selenium import webdriver
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from tkinter import *
import win32com.client
import win32com.client as win32
import os


ie_options = webdriver.IeOptions()
ie_options.attach_to_edge_chrome = True
ie_options.edge_executable_path = "C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe"

linha = 1

try:
    workbook = load_workbook("base.xlsx")
except:
    messagebox.showinfo(title="Accionar Mac Gestor", message="Base corrompida!")
    exit()
    
worksheet = workbook.active

def login():

    Login = "san.01125"

    arq = open(r"\\192.168.0.9\Publica\### PUBLICA - MAC BARBOSA ###\00.SITES\01.SANTANDER_FINANCEIRA\GESTOR.txt","r")

    linhas = arq.readlines()

    for linha in linhas:
        Senha = linha

    browser.get("https://www.gestaojudicial.com.br/Paginas/Principal/_FSet_Abertura.asp")

    elementLogin = browser.find_element(By.ID,"txtcd_Logon")

    elementSenha = browser.find_element(By.ID,"txtcd_Pwd")

    elementSeg = browser.find_element(By.ID,"CodSegInformado")

    button = browser.find_element(By.ID,"btOK")

    elementSegCriado = str(browser.find_element(By.ID,"CodSegCriado").get_attribute("value"))

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementLogin, Login)

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementSenha, Senha)

    browser.execute_script('arguments[0].setAttribute("value", arguments[1])', elementSeg, elementSegCriado)

    browser.execute_script('arguments[0].click()', button)

    time.sleep(2)

    browser.switch_to.frame('FraMenu')

    buttonPJ = browser.find_element(By.ID,"PJTxt")

    browser.execute_script('arguments[0].click()', buttonPJ)

def pesquisaCausa():

    time.sleep(1)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraMenu')

    pesqCausa = browser.find_element(By.ID,"PCTxt")

    browser.execute_script('arguments[0].click()', pesqCausa)

    browser.switch_to.default_content()

def inserirContrato(): 

    global linha

    time.sleep(4)

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('fraConsulta')

    browser.switch_to.frame('fraPesquisa')

    inputPesq = browser.find_element(By.ID,"txtPesquisa")

    contrato = worksheet.cell(row=linha + 1, column=1).value

    browser.execute_script("arguments[0].setAttribute('value', arguments[1])", inputPesq, contrato)

    button = browser.find_element(By.ID,"btnPesquisar")

    browser.execute_script("arguments[0].click()", button)

def selecionaCausa():

    global worksheet, linha, workbook

    time.sleep(1)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('fraConsulta')

    browser.switch_to.frame('fraResultado')

    tabelas = browser.find_elements(By.TAG_NAME,"tbody")

    linhas = tabelas[1].find_elements(By.TAG_NAME,"tr")

    time.sleep(2)

    quantidadeLinhas = 0

    arrayIndice = []

    for linhaFor in linhas:

        quantidadeLinhas += 1

        if linhaFor.value_of_css_property('background-Color') == "white":
            arrayIndice.append(quantidadeLinhas)
            
            
    if len(arrayIndice) == 0:

        worksheet.cell(row= linha + 1,column=3).value = "Parte contraria não encontrada"

        workbook.save("base.xlsx")

        workbook.save("backup.xlsx")

        browser.quit()

        iniciar()

    elif len(arrayIndice) == 1:

        colunas = linhas[arrayIndice[0] - 1].find_elements(By.TAG_NAME,"td")

    elif len(arrayIndice) > 1:

        colunas = linhas[arrayIndice[-1] - 1].find_elements(By.TAG_NAME,"td")

    checkbox = colunas[5].find_element(By.TAG_NAME,"input")

    browser.execute_script("arguments[0].click()", checkbox)

    browser.execute_script("arguments[0].checked = true", checkbox)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('fraConsulta')

    browser.switch_to.frame('fraPesquisa')

    img = browser.find_elements(By.TAG_NAME,"img")

    browser.execute_script("arguments[0].click()", img[4])

    time.sleep(2) 

    browser.switch_to.default_content()

def entrarSaldo():

    time.sleep(2)

    global slinhas

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraFichas')
    
    linkSaldo = browser.find_element(By.NAME,"CobSalTxt")

    browser.execute_script("arguments[0].click()", linkSaldo)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraVazio')

    planilhaCalc = browser.find_element(By.NAME,"btnPlanilha")

    browser.execute_script("arguments[0].click()", planilhaCalc)

    time.sleep(2)

    planilha = browser.find_element(By.NAME,"cboPlanilha")

    browser.execute_script("arguments[0].setAttribute('value', 'A')", planilha)

    Atualizado = browser.find_element(By.ID,"Radio2")

    browser.execute_script("arguments[0].click()", Atualizado)
    
    btnConfirmar = browser.find_element(By.NAME,"btnCalcular")

    browser.execute_script("arguments[0].click()", btnConfirmar)

    time.sleep(2)

    browser.switch_to.frame('fraPrice')

    stabelas = browser.find_element(By.ID,"tblSumario")

    slinhas = stabelas.find_elements(By.TAG_NAME,"tr")

    texto = slinhas[2].text

    valor = texto.replace("Total do Débito :\n", "")

    valor = float(valor.replace(".","").replace(",","."))

    if valor > 30000:
        worksheet.cell(row= linha + 1,column=4).value = "NÃO"

    elif valor < 30000:    
        worksheet.cell(row= linha + 1,column=4).value = "SIM"

    if valor > 10000:
        worksheet.cell(row= linha + 1,column=5).value = "NÃO"

    elif valor < 10000:
        worksheet.cell(row= linha + 1,column=5).value = "SIM"   

    worksheet.cell(row= linha + 1,column=6).value = valor

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('fraVazio')

    btnExportar = browser.find_element(By.NAME,"btnExportar")

    browser.execute_script("arguments[0].click()", btnExportar)

    usuario = os.environ.get("USERNAME")

    localDownload = "C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xls"

    verificaDownload(localDownload)

    time.sleep(3)

    salvaArquivo(worksheet.cell(row= linha +1,column=1).value, usuario)

    worksheet.cell(row= linha + 1,column=7).value = "OK"

def all_ok(hwnd, param):
        child_handles.append(hwnd)

def verificaDownload(localDownload):

    global segue_lopp, child_handles
    
    hwnd = win32gui.FindWindow('Chrome_WidgetWin_1',  """Gestor Jurídico e-Xyon - Perfil 1 — Microsoft​ Edge""")
            
    hwnd2= 0

    child_handles = []

    cont = True

    FrameBar = 0

    while cont == True:  

        time.sleep(7)

        win32gui.EnumChildWindows(hwnd, all_ok, None)

        for array in child_handles:
            
            if win32gui.GetWindowText(array) == 'Gestor Jurídico e-Xyon - Internet Explorer':
                hwnd2 = array
                break
            elif win32gui.GetClassName(array) == 'Frame Notification Bar':
                FrameBar = array
                break
        
        msgFWebpage = win32gui.FindWindow(None ,"Mensagem da página da web") #Message from webpage Mensagem da página da web
        
        btn = win32gui.FindWindowEx(msgFWebpage, 0, "Button", 'OK')
        
        vbScript = win32gui.FindWindow(None, 'VBScript: Atenção')
        
        btnVBScript = win32gui.FindWindowEx(vbScript, 0, "Button", 'OK')

        if FrameBar > 0:

            #clicking the down arrow next to save
            SaveBtn = uia.SplitButtonControl(SearchDepth=5, Name = '6')

                #wrap up the Invoke statement to try until the box shows up
            while True:
                try:
                    SaveBtn.GetInvokePattern().Invoke()
                except LookupError:
                    time.sleep(0.5)
                else:
                    break
                
            time.sleep(2)
            
                #Find/map saveas button and click the saveas option
            SaveasBtn = uia.MenuItemControl(SearchDepth=2, Name = 'Salvar como') # Salvar como
            SaveasBtn.GetInvokePattern().Invoke()
            
            time.sleep(1)
            
                #Find Save as dialog and click on address bar
            Savedlg = uia.WindowControl(SearchDepth=1, Name = 'Salvar como') #Save As Salvar como

            time.sleep(1)
            
                #Name File
            FileName = Savedlg.EditControl(SearchDepth=7, Name='Nome:')# Nome: File name:
            FileName.SendKeys(localDownload)

            #Click Saveas Dlg Save Button
            SaveBtn = Savedlg.ButtonControl(SearchDepth=2, Name="Salvar") #Save Salvar
            SaveBtn.GetInvokePattern().Invoke()

            Savedlgf = True
    
            while Savedlgf == True: 
                try:    
                    Savedlg = uia.WindowControl(SearchDepth=1, Name = 'Salvar como') #Save As Salvar como
                    print(Savedlg)
                    time.sleep(5)
                except:
                    print('EXCEPT') 
                    Savedlgf = False

            cont = False

        elif btnVBScript > 0:
            
            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)
            
            win32api.SendMessage(btnVBScript, win32con.BM_CLICK, 0, 0)
            
            time.sleep(2)                    

            cont = False

            browser.quit()

def salvaArquivo(contrato,usuario):

    global localSalvePDF

    if cluster == "SIM":

        localSalvePDF = "\\\\192.168.0.9\\Operacional\\3- LOCALIZAÇÃO – Resp. Bruno\\01 - RASTREADOR\\PUXADA\\" + str(contrato)

    elif cluster == "NÃO":

        localSalvePDF = "\\\\192.168.0.9\\Juridico\\3. PRÉ-JURÍCIDO – Resp. Simone\\PUXADA 2016\\" + str(contrato) 

    fname = "C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xls"
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    excel.Application.Quit()

    workbook = load_workbook("C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xlsx")

    worksheet = workbook.active

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = 9
    worksheet.page_margins.left = 0.01
    worksheet.page_margins.right = 0.01
    worksheet.page_margins.top = 0.01
    worksheet.page_margins.bottom = 0.01
    worksheet.page_margins.header = 0.01
    worksheet.page_margins.footer = 0.01

    workbook.save("C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xlsx")

    workbook.close()

    o = win32com.client.Dispatch("Excel.Application")

    o.Visible = False

    wb_path = "C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xlsx"

    wb = o.Workbooks.Open(wb_path)

    if not os.path.exists(localSalvePDF):
        os.makedirs(localSalvePDF)

    path_to_pdf = localSalvePDF + "\\PLANILHA DEBITO.pdf"

    time.sleep(0.1)

    wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)

    time.sleep(0.1)

    o.Quit()

    time.sleep(0.1)

    time.sleep(0.1)
    os.remove("C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xls")
    time.sleep(0.1)
    os.remove("C:\\Users\\"+ str(usuario) +"\\Downloads\\planilhaDeDebito.xlsx")
    time.sleep(0.1)

def verificaCadeado():
    
    time.sleep(3)

    browser.switch_to.default_content()

    browser.switch_to.frame('FraDetalhe')

    browser.switch_to.frame('FraVazio')

    browser.switch_to.frame('l_fraContent')

    browser.switch_to.frame('Frame_Processo')
    
    divs = browser.find_elements(By.TAG_NAME,"div")

    img =  divs[2].find_element(By.TAG_NAME,"img")

    if "leaf_lock.gif" in img.get_attribute('src'):
            
            worksheet.cell(row= linha + 1,column=3).value = "FECHADO"

            links = divs[2].find_element(By.TAG_NAME,"a")

            browser.execute_script("arguments[0].click()", links)

            entrarSaldo()

    elif "leaf_open.gif" in img.get_attribute('src'):

            worksheet.cell(row= linha + 1,column=3).value = "ABERTO"

            links = divs[2].find_element(By.TAG_NAME,"a")

            browser.execute_script("arguments[0].click()", links)

            entrarSaldo()

def iniciar():

    global linha, browser, cluster

    # browser = webdriver.Ie()
    browser = webdriver.Ie(options=ie_options)

    login()

    while worksheet.cell(row= linha + 1,column=1).value != None:
        
        if worksheet.cell(row= linha + 1,column=7).value == None:

            cluster = worksheet.cell(row=linha + 1, column=2).value
            
            pesquisaCausa()

            inserirContrato()

            selecionaCausa()

            verificaCadeado()

            linha = linha + 1

            workbook.save("base.xlsx")
            
            workbook.save("backup.xlsx")

        else:
            linha = linha + 1

try:
    
    iniciar()
    
    messagebox.showinfo(title="Mac Gestor", message="Fim")
    

except:

    workbook.save("base.xlsx")

    browser.quit()

    iniciar()