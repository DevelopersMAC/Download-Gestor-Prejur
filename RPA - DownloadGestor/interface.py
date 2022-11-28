from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from tkinter import *
from os import system


def set_text():
    try:
        workbook = load_workbook(filename="base.xlsx")

        worksheet = workbook.active

        row = 1
        row2 = 1

        while worksheet.cell(row= row, column=1).value != None:
            row += 1   

        while worksheet.cell(row= row2, column=6).value != None:
            row2 += 1   

        row -= 2
        row2 -= 2
        
        if row == 0:
            return messagebox.showwarning(title="Accionar Mac Gestor Pre-juridico", message="Base sem contratos.")

        NrLabel.configure(state="normal")
        NrLabel.delete(0,END)
        NrLabel.insert(0,str(row))
        NrLabel.configure(state="disabled")
        
        NrOKLabel.configure(state="normal")
        NrOKLabel.delete(0,END)
        NrOKLabel.insert(0,str(row2))
        NrOKLabel.configure(state="disabled")
    
        return
    except:
        return messagebox.showerror(title="Accionar Mac Gestor Pre-juridico", message="Não foi possivel encontrar a base.\nFavor Criar a base.\nE inserir a base na mesma.")

def criarBase():
    try:
        workbook = load_workbook("base.xlsx")

        return messagebox.showwarning(title="Accionar Mac Gestor Pre-juridico", message="ATENÇÃO!\nJa existe um arquivo base!") 

    except:
        workbook = Workbook(write_only = False)

        worksheet = workbook.active
        
        worksheet.cell(row=1,column=1).value = "CONTRATO"

        worksheet.cell(row=1,column=2).value = "CLUSTER"

        worksheet.cell(row=1,column=3).value = "CADEADO"

        worksheet.cell(row=1,column=4).value = "INFRIOR A 30 MIL"

        worksheet.cell(row=1,column=5).value = "INFRIOR A 10 MIL"

        worksheet.cell(row=1,column=6).value = "VALOR"

        worksheet.cell(row=1,column=7).value = "ARQUIVO"

        workbook.save("base.xlsx")

        return messagebox.showinfo(title="Accionar Mac Gestor Pre-juridico", message="Base criada com sucesso!\nFavor insira os contratos na base!") 

def iniciar():
    try:
        system("main.exe")
    except:
        return messagebox.showwarning(title="Accionar Mac Gestor Pre-juridico", message="Sem base\nPor favor Crie a base e\nalimente a base.")

def descorromper():
    try:
        workbook = load_workbook("backup.xlsx")

        workbook.save("base.xlsx")
        
        workbook = load_workbook("backup.xlsx")

        return messagebox.showinfo(title="Accionar Mac Gestor Pre-juridico", message="Base Reparada!")
    except:
        return messagebox.showerror(title="Accionar Mac Gestor Pre-juridico", message="Erro ao Reparar\nbase.")

def telaPrincipal():

    global NrLabel,NrOKLabel

    janela = Tk()

    janela.title("Accionar Mac Gestor Pre-juridico")
    janela.geometry("470x140")
    janela.configure(background="WHITE")
    janela.resizable(width=False, height=False)
    janela.attributes("-alpha", 1)
    janela.iconbitmap("LogoM.ico")
    
    logo = PhotoImage(file="editar.png")

    leftFrame = Frame(janela, width=200, height=200, bg="#04dc9c", relief="raise")
    leftFrame.pack(side=LEFT)

    RightFrame = Frame(janela, width=267 , height=200, bg="#b4f236", relief="raise")
    RightFrame.pack(side=RIGHT)

    LogoLabel = Label(leftFrame, image=logo, bg="#04dc9c")
    LogoLabel.place(x=25, y= -6)

    TotalContratoLabel = Entry(RightFrame, width=13, font=("Century Gothic", 16), fg="Black")
    TotalContratoLabel.insert(0, "Total contratos:")
    TotalContratoLabel.configure(state="disabled")
    TotalContratoLabel.place(x=10, y=10)

    NrLabel = Entry(RightFrame, width=5, font=("Century Gothic", 16), fg="Black")
    NrLabel.insert(0, "0")
    NrLabel.configure(state="disabled")
    NrLabel.place(x=180, y=10)

    ContratoOkLabel = Entry(RightFrame, width=13, font=("Century Gothic", 16), fg="Black")
    ContratoOkLabel.insert(0, "Contratos OK:")
    ContratoOkLabel.configure(state="disabled")
    ContratoOkLabel.place(x=10, y=45)

    NrOKLabel = Entry(RightFrame, width=5, font=("Century Gothic", 16), fg="Black")
    NrOKLabel.insert(0, "0")
    NrOKLabel.configure(state="disabled")
    NrOKLabel.place(x=180, y=45)

    b1 = Button(RightFrame,width= 7,text="Calular\ncontratos",command=lambda:set_text())
    b1.place(x=5, y=90)

    b2 = Button(RightFrame,width= 7,text="Criar\nBase",command=criarBase)
    b2.place(x=65, y=90)

    b3 = Button(RightFrame,width= 7,text="Executar\nRPA",command=iniciar)
    b3.place(x=185, y=90)
    
    b4 = Button(RightFrame,width= 7,text="Reparar\nbase",command=descorromper)
    b4.place(x=125, y=90)

    janela.mainloop()

telaPrincipal()